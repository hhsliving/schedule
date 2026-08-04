# -*- coding: utf-8 -*-
"""weeks/ 안의 편성표와 비중 파일을 자동으로 짝짓습니다.

동일한 주차의 편성표가 여러 개 있으면 가장 최근 커밋의 편성표만 사용합니다.

비중 파일은 다음 순서로 선택합니다.

1. 선택된 최신 편성표와 같은 커밋에 올라온 비중
2. 기존 .pairs.json에 연결되어 있던 비중 중 가장 최근 세트
3. Git 이력이 없는 로컬 환경에서는 파일 수정 시각 기준

오래된 동일 주차 편성표와 오래된 비중 파일은 삭제하지 않아도 자동으로 무시됩니다.
"""

import glob
import json
import os
import re
import subprocess

import openpyxl

import read_ratio_xlsx


SHEET = "(TV)편성기획"

TITLE = re.compile(
    r"◆\s*(\d{4})년\s*(\d{2})월\s*(\d+)주\s*편성표\s*"
    r"\((\d{2})/(\d{2})\s*~\s*(\d{2})/(\d{2})\)"
)


def dup_suffix(path):
    """파일명 끝의 (n)을 반환합니다.

    예:
        undefined_20260805.xlsx     → 0
        undefined_20260805 (1).xlsx → 1

    편성표 제목의 (0817 ~ 0823) 같은 날짜 괄호는 오인하지 않습니다.
    """

    base = os.path.splitext(os.path.basename(path))[0]
    match = re.search(r"\((\d+)\)\s*$", base)

    return int(match.group(1)) if match else 0


def classify(path):
    """엑셀 파일이 편성표인지 비중 리포트인지 판별합니다."""

    try:
        workbook = openpyxl.load_workbook(
            path,
            read_only=True,
            data_only=False,
        )
        sheet_names = workbook.sheetnames
        workbook.close()

    except Exception:
        return None

    if SHEET in sheet_names:
        return "sched"

    try:
        read_ratio_xlsx.read(path)
        return "ratio"

    except Exception:
        return None


def sched_key(path):
    """편성표 A1 제목에서 주차 정보를 읽습니다."""

    workbook = openpyxl.load_workbook(
        path,
        read_only=True,
        data_only=False,
    )

    title = str(
        workbook[SHEET]["A1"].value or ""
    ).strip()

    workbook.close()

    match = TITLE.match(title)

    if not match:
        return None, None

    year = int(match.group(1))
    month = int(match.group(2))
    week = int(match.group(3))

    start = (
        f"{match.group(4)}/{match.group(5)}"
    )
    end = (
        f"{match.group(6)}/{match.group(7)}"
    )

    cross = match.group(4) != match.group(6)

    key = f"{year}-{month}-{week}"

    info = {
        "y": year,
        "m": month,
        "w": week,
        "span": f"{start} ~ {end}",
        "cross": cross,
    }

    return key, info


def commit_info(path, root):
    """파일이 마지막으로 변경된 Git 커밋 정보를 반환합니다.

    반환값:
        commit_hash
        commit_time
        commit_rank

    commit_rank는 저장소 역사상 뒤쪽 커밋일수록 큰 숫자입니다.
    """

    try:
        relative_path = os.path.relpath(path, root)

        result = subprocess.run(
            [
                "git",
                "-C",
                root,
                "log",
                "-1",
                "--format=%H\t%ct",
                "--",
                relative_path,
            ],
            capture_output=True,
            text=True,
            timeout=15,
            check=False,
        )

        value = result.stdout.strip()

        if not value:
            return None, 0, 0

        commit_hash, commit_time = value.split(
            "\t",
            1,
        )

        rank_result = subprocess.run(
            [
                "git",
                "-C",
                root,
                "rev-list",
                "--count",
                commit_hash,
            ],
            capture_output=True,
            text=True,
            timeout=15,
            check=False,
        )

        rank_text = rank_result.stdout.strip()

        commit_rank = (
            int(rank_text)
            if rank_text.isdigit()
            else 0
        )

        return (
            commit_hash,
            int(commit_time),
            commit_rank,
        )

    except Exception:
        return None, 0, 0


def freshness(item):
    """파일의 최신 여부를 비교하기 위한 정렬 키입니다."""

    # 로컬에서 아직 커밋하지 않은 파일은 기존 커밋 파일보다
    # 최신 파일로 취급합니다.
    if not item.get("commit"):
        return (
            1,
            item.get("mtime", 0),
            0,
            os.path.basename(item["file"]),
        )

    return (
        0,
        item.get("commit_rank", 0),
        item.get("commit_time", 0),
        os.path.basename(item["file"]),
    )


def load_saved_pairs(store):
    """기존 .pairs.json을 읽습니다."""

    if not os.path.exists(store):
        return {}

    try:
        with open(
            store,
            encoding="utf-8",
        ) as file:
            value = json.load(file)

        return value if isinstance(value, dict) else {}

    except Exception:
        return {}


def select_latest_schedules(schedules):
    """동일 주차 편성표 중 최신 파일 하나만 선택합니다."""

    grouped = {}

    for schedule in schedules:
        grouped.setdefault(
            schedule["key"],
            [],
        ).append(schedule)

    selected = {}

    for key, candidates in grouped.items():
        chosen = max(
            candidates,
            key=freshness,
        )

        selected[key] = chosen

        if len(candidates) > 1:
            ignored = [
                os.path.basename(item["file"])
                for item in candidates
                if item["file"] != chosen["file"]
            ]

            print(
                f"  ↻ {key}: 최신 편성표 선택"
            )
            print(
                f"      사용: {os.path.basename(chosen['file'])}"
            )

            for filename in ignored:
                print(
                    f"      무시: {filename}"
                )

    return selected


def assign_ratio_key(
    ratio,
    selected_schedules,
    saved_pairs,
    weeks_dir,
):
    """비중 파일이 어느 주차에 해당하는지 결정합니다."""

    relative_path = os.path.relpath(
        ratio["file"],
        weeks_dir,
    )

    # 1. 최신 편성표와 같은 커밋에 올라온 비중
    if ratio.get("commit"):
        same_commit = [
            schedule
            for schedule in selected_schedules.values()
            if schedule.get("commit") == ratio["commit"]
        ]

        if len(same_commit) == 1:
            return same_commit[0]["key"]

        if len(same_commit) > 1:
            # 한 커밋에 여러 주차 편성표가 있다면
            # 파일 수정 시각이 가장 가까운 편성표를 선택합니다.
            nearest = min(
                same_commit,
                key=lambda schedule: abs(
                    schedule["mtime"] - ratio["mtime"]
                ),
            )

            print(
                f"  ! {os.path.basename(ratio['file'])}: "
                "같은 커밋에 편성표가 여러 개입니다 "
                f"→ {nearest['key']} 선택"
            )

            return nearest["key"]

    # 2. 기존 저장 연결
    saved_key = saved_pairs.get(relative_path)

    if saved_key in selected_schedules:
        return saved_key

    # 3. Git 이력이 없는 로컬 실행 시에만 mtime 사용
    if not ratio.get("commit") and selected_schedules:
        nearest = min(
            selected_schedules.values(),
            key=lambda schedule: abs(
                schedule["mtime"] - ratio["mtime"]
            ),
        )

        if abs(
            nearest["mtime"] - ratio["mtime"]
        ) < 120:
            return nearest["key"]

    return None


def select_latest_ratio_set(
    schedule,
    ratio_records,
):
    """해당 주차에 연결된 비중 중 최신 세트만 선택합니다."""

    if not ratio_records:
        return []

    schedule_commit = schedule.get("commit")

    # 최신 편성표와 같은 커밋의 비중이 있으면
    # 다른 모든 비중은 무시합니다.
    if schedule_commit:
        same_commit = [
            ratio
            for ratio in ratio_records
            if ratio.get("commit") == schedule_commit
        ]

        if same_commit:
            selected = same_commit

        else:
            selected = None

    else:
        selected = None

    # 같은 커밋 비중이 없으면 연결된 비중 중
    # 가장 최근 커밋에 들어 있는 세트를 사용합니다.
    if selected is None:
        committed = [
            ratio
            for ratio in ratio_records
            if ratio.get("commit")
        ]

        if committed:
            newest = max(
                committed,
                key=freshness,
            )

            newest_commit = newest["commit"]

            selected = [
                ratio
                for ratio in committed
                if ratio["commit"] == newest_commit
            ]

        else:
            # Git 없는 로컬 환경
            newest = max(
                ratio_records,
                key=freshness,
            )

            selected = [
                ratio
                for ratio in ratio_records
                if abs(
                    ratio["mtime"] - newest["mtime"]
                ) < 120
            ]

    selected.sort(
        key=lambda ratio: (
            dup_suffix(ratio["file"]),
            os.path.basename(ratio["file"]),
        )
    )

    selected_paths = {
        ratio["file"]
        for ratio in selected
    }

    ignored = [
        ratio
        for ratio in ratio_records
        if ratio["file"] not in selected_paths
    ]

    if ignored:
        print(
            f"  ↻ {schedule['key']}: "
            f"이전 비중 {len(ignored)}개 무시"
        )

        for ratio in ignored:
            print(
                f"      무시: "
                f"{os.path.basename(ratio['file'])}"
            )

    return selected


def pair(weeks_dir, root):
    """편성표와 비중 파일을 최신 주차 세트 기준으로 반환합니다."""

    files = [
        path
        for path in glob.glob(
            os.path.join(
                weeks_dir,
                "*.xlsx",
            )
        )
        if not os.path.basename(path).startswith("~$")
    ]

    schedules = []
    ratios = []

    for path in files:
        file_type = classify(path)

        commit_hash, commit_time, commit_rank = (
            commit_info(path, root)
        )

        common = {
            "file": path,
            "commit": commit_hash,
            "commit_time": commit_time,
            "commit_rank": commit_rank,
            "mtime": os.path.getmtime(path),
        }

        if file_type == "sched":
            key, info = sched_key(path)

            if not key:
                print(
                    f"  ! {os.path.basename(path)}: "
                    "편성표 A1 제목을 읽지 못해 무시"
                )
                continue

            schedules.append(
                {
                    **common,
                    "key": key,
                    "info": info,
                }
            )

        elif file_type == "ratio":
            ratios.append(common)

        else:
            print(
                f"  ! {os.path.basename(path)}: "
                "편성표도 비중도 아니므로 무시"
            )

    # 동일 주차 중 최신 편성표 하나만 남깁니다.
    selected_schedules = select_latest_schedules(
        schedules
    )

    store = os.path.join(
        weeks_dir,
        ".pairs.json",
    )

    saved_pairs = load_saved_pairs(store)

    by_key = {}

    for key, schedule in selected_schedules.items():
        by_key[key] = {
            "file": schedule["file"],
            **schedule["info"],
            "key": key,
            "ratios": [],
            "_schedule": schedule,
            "_ratio_records": [],
        }

    # 모든 비중 파일을 우선 주차에 배정합니다.
    for ratio in ratios:
        key = assign_ratio_key(
            ratio=ratio,
            selected_schedules=selected_schedules,
            saved_pairs=saved_pairs,
            weeks_dir=weeks_dir,
        )

        if not key:
            print(
                f"  ! {os.path.basename(ratio['file'])}: "
                "어느 편성표의 비중인지 알 수 없어 무시"
            )
            continue

        if key not in by_key:
            print(
                f"  ! {os.path.basename(ratio['file'])}: "
                f"배정된 주차 {key}의 편성표가 없어 무시"
            )
            continue

        by_key[key]["_ratio_records"].append(
            ratio
        )

    # 각 주차별 최신 비중 세트만 남깁니다.
    new_saved_pairs = {}

    for key, item in by_key.items():
        selected_ratios = select_latest_ratio_set(
            schedule=item["_schedule"],
            ratio_records=item["_ratio_records"],
        )

        item["ratios"] = [
            ratio["file"]
            for ratio in selected_ratios
        ]

        for ratio in selected_ratios:
            relative_path = os.path.relpath(
                ratio["file"],
                weeks_dir,
            )

            new_saved_pairs[relative_path] = key

        del item["_schedule"]
        del item["_ratio_records"]

    # 오래된 연결은 제거하고 현재 사용 중인 비중만 기록합니다.
    try:
        with open(
            store,
            "w",
            encoding="utf-8",
        ) as file:
            json.dump(
                new_saved_pairs,
                file,
                ensure_ascii=False,
                indent=1,
            )

    except Exception as exception:
        print(
            f"  ! .pairs.json 저장 실패: {exception}"
        )

    return sorted(
        by_key.values(),
        key=lambda item: (
            item["y"],
            item["m"],
            item["w"],
        ),
    )


if __name__ == "__main__":
    root = os.path.dirname(
        os.path.dirname(
            os.path.abspath(__file__)
        )
    )

    weeks_dir = os.path.join(
        root,
        "weeks",
    )

    for item in pair(
        weeks_dir,
        root,
    ):
        ratio_names = " + ".join(
            os.path.basename(path)
            for path in item["ratios"]
        ) or "(비중 없음)"

        print(
            f"  {item['key']:10s} "
            f"{item['span']:15s} "
            f"{'월교차' if item['cross'] else '      '} "
            f"← {ratio_names}"
        )
