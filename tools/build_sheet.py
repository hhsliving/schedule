# -*- coding: utf-8 -*-
"""
VBA 매크로 '편성_사업부_팀_동시실행_2025_Final_v3' 를 openpyxl로 한 줄씩 재현.

[A] 원본 (TV)편성기획 시트 (사업부용):
  1. A1:AK3 병합(가운데), AL1:AQ3 병합 + 고정 라벨 수식(bold)
  2. 5~6행 삭제
  3. 테두리: A4:AQ30 thin 그리드 + 내부색 초기화 → medium 박스들
     (A1:AQ30 / A4:AQ4 / B4:G30 / N4:S30 / Z4:AE30 / AL4:AQ30 / A1:AQ30 / A4:AQ5)
  4. 색칠: 'hit' 포함 셀 자체(텍스트 유지, ReplaceFormat 트릭) →
     '가전'(노랑)/'리빙'(주황)/'주방'(연두): 셀+왼쪽1~3+오른쪽1 →
     'hit & best': 왼쪽1~3
  5. Sheet1 요약표(A1:C5) 생성 + 비중값 입력
  6. Rows 3:4 삽입(제목영역 5행으로 확장) → Sheet1 표 캡처 이미지를 AL1에 삽입
     (크기: VBA H92.69*1.09pt, W170.08*1.04pt)

[B] 팀편성 복사본:
  - 동일 서식/테두리, 단 AL1:AQ3 수식은 월교차 분기(Sheet1 비중 참조),
    색칠은 '가전' + 'hit & best'만, AL1:AQ3 내부색 초기화,
    사업부와 요일 행 높이를 맞추기 위한 상단 2행 삽입, 그림 없음
"""
import re, os
from copy import copy
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from PIL import Image as PILImage, ImageDraw, ImageFont

SRC = os.environ.get("SRCFILE")
if not SRC:
    raise SystemExit("SRCFILE 환경변수에 RAW 편성표 경로를 지정하세요")
OUT = os.environ.get("OUTFILE", "out.xlsx")

MAX_COL = 43  # AQ

thin = Side(style="thin", color="000000")
medium = Side(style="medium", color="000000")
thick = Side(style="thick", color="000000")
NO_FILL = PatternFill(fill_type=None)

def solid(hex6):
    return PatternFill(start_color=hex6, end_color=hex6, fill_type="solid")

YELLOW = "FFFF00"   # 가전
ORANGE = "FFC000"   # 리빙
GREEN  = "C6E0B4"   # 주방
SKY    = "ADD8E6"   # hit & best
HEADER_BLUE = "DDEBF7"

# 사용자 제공 비중값
BLANK = os.environ.get("BLANK", "") == "1"
VAL_B = {"가전": None, "리빙": None, "주방": None} if BLANK else eval(os.environ.get("VB", '{"가전":0.0410,"리빙":0.0713,"주방":0.0674}'))
VAL_C = {"가전": None, "리빙": None, "주방": None} if BLANK else eval(os.environ.get("VC", '{"가전":0.0449,"리빙":0.0656,"주방":0.0656}'))

# ---------------------------------------------------------------
# 헬퍼
# ---------------------------------------------------------------
def unmerge_intersecting(ws, r1, r2, c1, c2):
    """지정 영역과 겹치는 기존 병합을 전부 해제 (VBA UnMerge 대응)."""
    for m in list(ws.merged_cells.ranges):
        if not (m.max_row < r1 or m.min_row > r2 or m.max_col < c1 or m.min_col > c2):
            ws.unmerge_cells(str(m))

def apply_box(ws, r1, r2, c1, c2, edge, inside_v="keep", inside_h="keep"):
    """VBA의 Borders(xlEdge*/xlInside*) 블록 대응.
    inside_v / inside_h: "keep"=기존 유지, None=제거, Side=지정"""
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            b = cell.border
            left, right, top, bottom = b.left, b.right, b.top, b.bottom
            if c == c1:
                left = edge
            elif inside_v != "keep":
                left = inside_v
            if c == c2:
                right = edge
            elif inside_v != "keep":
                right = inside_v
            if r == r1:
                top = edge
            elif inside_h != "keep":
                top = inside_h
            if r == r2:
                bottom = edge
            elif inside_h != "keep":
                bottom = inside_h
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

def thin_grid_and_clear(ws):
    """VBA: A4:AQ30 전체 thin 테두리 + Interior.Pattern = xlNone (색 초기화)"""
    grid = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(4, 31):
        for c in range(1, MAX_COL + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = grid
            cell.fill = NO_FILL

def full_border_scheme(ws):
    """VBA 섹션2의 medium 박스 순서 그대로."""
    thin_grid_and_clear(ws)                                   # A4:AQ30 thin + 색초기화
    apply_box(ws, 1, 30, 1, MAX_COL, thick)                   # A1:AQ30 바깥쪽 굵게
    apply_box(ws, 4, 4, 1, MAX_COL, medium, inside_v=thin,
              inside_h=None)                                  # A4:AQ4
    apply_box(ws, 4, 30, 2, 7, medium, inside_v=thin)         # B4:G30 (월)
    apply_box(ws, 4, 30, 14, 19, medium, inside_v=thin)       # N4:S30 (수)
    apply_box(ws, 4, 30, 26, 31, medium, inside_v=thin)       # Z4:AE30 (금)
    apply_box(ws, 4, 30, 38, MAX_COL, medium, inside_v=thin)  # AL4:AQ30 (일)
    apply_box(ws, 1, 30, 1, MAX_COL, thick)                   # A1:AQ30 (재적용 바깥쪽 굵게)

def color_hit_self(ws):
    """VBA: Cells.Replace What:="hit", Replacement:="", ReplaceFormat 트릭
    → 'hit' 포함 셀에 하늘색 적용, 텍스트는 유지."""
    fill = solid(SKY)
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and "hit" in v.lower():
                cell.fill = fill

def left_offsets(col):
    if col >= 4:
        return [-1, -2, -3]
    if col == 3:
        return [-1, -2]
    if col == 2:
        return [-1]
    return []

def color_depts(ws, dept_colors):
    """'가전'/'리빙'/'주방' 셀: 자기자신 + 왼쪽1~3 + 오른쪽1"""
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if not isinstance(v, str):
                continue
            key = v.strip().strip("'").strip()
            if key in dept_colors:
                fill = solid(dept_colors[key])
                col = cell.column
                cell.fill = fill
                for off in left_offsets(col):
                    ws.cell(row=cell.row, column=col + off).fill = fill
                if col < MAX_COL + 1:
                    ws.cell(row=cell.row, column=col + 1).fill = fill

def color_hitbest_left(ws):
    """'hit & best' 셀: 왼쪽1~3만 하늘색 (셀 자체는 ReplaceFormat에서 이미 처리)"""
    fill = solid(SKY)
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and "hit & best" in v.lower():
                col = cell.column
                for off in left_offsets(col):
                    ws.cell(row=cell.row, column=col + off).fill = fill

def snapshot_heights(ws):
    return {r: d.height for r, d in ws.row_dimensions.items() if d.height is not None}

def rebuild_heights(ws, mapping):
    for r in list(ws.row_dimensions.keys()):
        del ws.row_dimensions[r]
    for r, h in mapping.items():
        ws.row_dimensions[r].height = h

def set_merge_border(ws, ref, left=None, right=None, top=None, bottom=None):
    """병합범위 둘레 테두리를 명시적으로 지정.
    openpyxl merge_cells()가 병합 시점에 anchor 테두리를 가장자리로 전파해
    기존 테두리를 파괴하므로, 병합 완료 후 이 함수로 복원한다."""
    c1, r1, c2, r2 = range_boundaries(ref)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            b = cell.border
            nb = {"left": b.left, "right": b.right, "top": b.top, "bottom": b.bottom}
            if c == c1 and left is not None:
                nb["left"] = left
            if c == c2 and right is not None:
                nb["right"] = right
            if r == r1 and top is not None:
                nb["top"] = top
            if r == r2 and bottom is not None:
                nb["bottom"] = bottom
            cell.border = Border(**nb)
    # anchor에도 병합 외곽 전체를 반영 (저장 시 전파 대비)
    anchor = ws.cell(row=r1, column=c1)
    anchor.border = Border(
        left=left if left is not None else anchor.border.left,
        right=right if right is not None else anchor.border.right,
        top=top if top is not None else anchor.border.top,
        bottom=bottom if bottom is not None else anchor.border.bottom,
    )

def finalize_merged_borders(ws, row_offset):
    """매크로 테두리 스펙 기준으로 상단 병합영역들의 둘레 테두리 확정.
    row_offset: wsTV는 3:4행 삽입으로 +2, wsCopy는 0."""
    t = 3 + row_offset   # 제목 병합 마지막 행
    h = 4 + row_offset   # 요일 헤더 행
    set_merge_border(ws, f"A1:AK{t}", left=thick, top=thick, right=thin)
    set_merge_border(ws, f"AL1:AQ{t}", left=thin, top=thick, right=thick)
    day_groups = [("B", "G", True), ("H", "M", False), ("N", "S", True),
                  ("T", "Y", False), ("Z", "AE", True), ("AF", "AK", False),
                  ("AL", "AQ", True)]
    for a, b, boxed in day_groups:
        side = medium if boxed else thin
        set_merge_border(ws, f"{a}{h}:{b}{h}",
                         left=side, right=side, top=medium, bottom=medium)

# ---------------------------------------------------------------
# 로드
# ---------------------------------------------------------------
wb = openpyxl.load_workbook(SRC)
wsTV = wb["(TV)편성기획"]

# 양식 통일: 과거 파일은 열 너비가 좁으므로 기준 파일(7월 양식)의 열 너비를 그대로 입힌다.
#            TEMPLATE 미지정이면 원본 너비 유지.
TEMPLATE = os.environ.get("TEMPLATE", "")
if TEMPLATE:
    _tw = openpyxl.load_workbook(TEMPLATE)
    _ts = _tw["(TV)편성기획"]
    _n = 0
    for _c in range(1, MAX_COL + 1):
        _L = get_column_letter(_c)
        _new = _ts.column_dimensions[_L].width
        if wsTV.column_dimensions[_L].width != _new:
            wsTV.column_dimensions[_L].width = _new
            _n += 1
    _tw.close()
    print(f"열 너비 통일: {_n}개 열 조정 (기준 {os.path.basename(TEMPLATE)})")

wsCopy = wb.copy_worksheet(wsTV)
wsCopy.title = wsTV.title + "-팀편성"

a1_text = wsTV["A1"].value or ""

# 월 교차 감지 (4행 '월 */*', '일 */*')
monMonth = sunMonth = None
for c in range(1, MAX_COL + 1):
    v = wsTV.cell(row=4, column=c).value
    if isinstance(v, str):
        s = v.strip()
        m = re.match(r"^월\s+(\d+)/\d+", s)
        if m:
            monMonth = int(m.group(1))
        m = re.match(r"^일\s+(\d+)/\d+", s)
        if m:
            sunMonth = int(m.group(1))
isCrossMonth = (monMonth and sunMonth and monMonth != sunMonth)
print("monMonth:", monMonth, "sunMonth:", sunMonth, "isCrossMonth:", isCrossMonth)

주차텍스트 = a1_text[8:14]  # VBA Mid(A1,9,6)

CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)

def process_sheet(ws, is_team_copy):
    heights0 = snapshot_heights(ws)

    # --- 섹션1: 상단 병합 ---
    unmerge_intersecting(ws, 1, 3, 1, 37)          # A1:AK3
    ws.merge_cells("A1:AK3")
    ws["A1"].alignment = CENTER_WRAP

    unmerge_intersecting(ws, 1, 3, 38, MAX_COL)    # AL1:AQ3
    ws.merge_cells("AL1:AQ3")

    if not is_team_copy:
        # [A] 고정 라벨 수식 (월교차 무관)
        ws["AL1"] = ('="가전팀"&CHAR(10)&MID(A1,9,6)&"차 편성 비중 :"'
                     '&CHAR(10)&MID(A1,9,3)&" 누적 편성 비중 :"')
    else:
        # [B] 월교차 분기 + Sheet1 비중 참조
        if isCrossMonth:
            ws["AL1"] = (f'="가전팀"&CHAR(10)'
                         f'&"{monMonth}월 누적 편성 비중 : "&TEXT(Sheet1!C3,"0.00%")'
                         f'&CHAR(10)&"{sunMonth}월 1주 편성 비중 : "&TEXT(Sheet1!B3,"0.00%")')
        else:
            ws["AL1"] = ('="가전팀"&CHAR(10)&MID(A1,9,6)&" 편성 비중 : "&TEXT(Sheet1!B3,"0.00%")'
                         '&CHAR(10)&MID(A1,9,3)&" 누적 편성 비중 : "&TEXT(Sheet1!C3,"0.00%")')
    ws["AL1"].font = Font(bold=True)
    ws["AL1"].alignment = CENTER_WRAP

    # --- 5~6행 삭제 ---
    ws.delete_rows(5, 2)
    # 행높이 재배치 (1~4 유지, 7+ → -2)
    new_h = {}
    for r, h in heights0.items():
        if r <= 4:
            new_h[r] = h
        elif r >= 7:
            new_h[r - 2] = h
    rebuild_heights(ws, new_h)

    # --- 섹션2: 테두리 (내부색 초기화 포함) ---
    full_border_scheme(ws)

    # --- 섹션3: 색칠 (매크로 순서: hit자체 → 부서 → hit&best 왼쪽) ---
    color_hit_self(ws)
    if is_team_copy:
        # [B] 는 ReplaceFormat 직후 AL1:AQ3 내부색 명시 초기화
        for r in range(1, 4):
            for c in range(38, MAX_COL + 1):
                ws.cell(row=r, column=c).fill = NO_FILL
        color_depts(ws, {"가전": YELLOW})
    else:
        color_depts(ws, {"가전": YELLOW, "리빙": ORANGE, "주방": GREEN})
    color_hitbest_left(ws)

    # --- 마지막: A4:AQ5 medium 박스 (헤더 2행) ---
    apply_box(ws, 4, 5, 1, MAX_COL, medium)
    apply_box(ws, 1, 30, 1, MAX_COL, thick) # 바깥쪽 테두리(thick) 덮어쓰기

process_sheet(wsTV, is_team_copy=False)
process_sheet(wsCopy, is_team_copy=True)

# ---------------------------------------------------------------
# 섹션4: Sheet1 요약표 + 비중값
# ---------------------------------------------------------------
if "Sheet1" in wb.sheetnames:
    wsCap = wb["Sheet1"]
else:
    wsCap = wb.create_sheet("Sheet1", 0)

for row in wsCap["A1:C5"]:
    for cell in row:
        cell.value = None
        cell.fill = NO_FILL

if isCrossMonth:
    wsCap["A1"] = ""
    wsCap["B1"] = f"{sunMonth}월 1주차"
    wsCap["C1"] = f"{monMonth}월 누적"
    wsCap["B1"].font = Font(bold=True, size=10)
    wsCap["C1"].font = Font(bold=True, size=10)
else:
    wsCap["A1"] = 주차텍스트
    wsCap["B1"] = "주"
    wsCap["C1"] = "월"
    wsCap["B1"].font = Font(bold=True)
    wsCap["C1"].font = Font(bold=True)
wsCap["A1"].font = Font(bold=True)
for c in ("A1", "B1", "C1"):
    wsCap[c].fill = solid(HEADER_BLUE)

wsCap["A2"] = "사업부"
if not BLANK:
    wsCap["B2"] = "=SUM(B3:B5)"
    wsCap["C2"] = "=SUM(C3:C5)"
wsCap["A3"] = "가전"
wsCap["A4"] = "리빙"
wsCap["A5"] = "주방"
for _r, _k in ((3, "가전"), (4, "리빙"), (5, "주방")):
    if VAL_B[_k] is not None: wsCap[f"B{_r}"] = VAL_B[_k]
    if VAL_C[_k] is not None: wsCap[f"C{_r}"] = VAL_C[_k]

wsCap["A3"].fill = solid(YELLOW)
wsCap["A4"].fill = solid(ORANGE)
wsCap["A5"].fill = solid(GREEN)
white = solid("FFFFFF")
for c in ("A2", "B2", "C2"):
    wsCap[c].fill = white
for r in range(3, 6):
    for col in ("B", "C"):
        wsCap[f"{col}{r}"].fill = white
for r in range(2, 6):
    for col in ("B", "C"):
        wsCap[f"{col}{r}"].number_format = "0.00%"

for row in wsCap["A1:C5"]:
    for cell in row:
        cell.alignment = CENTER_WRAP
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
apply_box(wsCap, 1, 5, 1, 3, thick)

# ---------------------------------------------------------------
# 섹션5: 두 시트의 상단 높이 통일
# 사업부/팀편성 모두 3:4행을 삽입해야 요일 행과 그 아래가 같은 위치에 놓인다.
# 사업부에만 Sheet1 표 이미지를 AL1에 붙인다.
# ---------------------------------------------------------------
def expand_header_to_five_rows(ws):
    """제목 영역을 3행에서 5행으로 늘리고 기존 요일/본문을 2행 내린다."""
    heights_before = snapshot_heights(ws)
    merges_before = [str(m) for m in ws.merged_cells.ranges]
    for merged in merges_before:
        ws.unmerge_cells(merged)

    ws.insert_rows(3, 2)

    # 병합 복원: 3행 이전 유지 / 3행 걸침 확장 / 3행 이후 +2
    for merged in merges_before:
        c1, r1, c2, r2 = range_boundaries(merged)
        if r2 < 3:
            nr1, nr2 = r1, r2
        elif r1 >= 3:
            nr1, nr2 = r1 + 2, r2 + 2
        else:
            nr1, nr2 = r1, r2 + 2
        ws.merge_cells(start_row=nr1, start_column=c1,
                       end_row=nr2, end_column=c2)

    # 삽입행(3,4)은 윗행(2행) 서식 상속 (Excel CopyOrigin 대응)
    for c in range(1, MAX_COL + 1):
        source = ws.cell(row=2, column=c)
        for r in (3, 4):
            target = ws.cell(row=r, column=c)
            target.border = copy(source.border)
            target.fill = copy(source.fill)

    # 행높이 재배치: 1~2 유지, 기존 3행 이후는 +2
    new_h = {}
    for r, h in heights_before.items():
        if r <= 2:
            new_h[r] = h
        else:
            new_h[r + 2] = h
    rebuild_heights(ws, new_h)

expand_header_to_five_rows(wsTV)
expand_header_to_five_rows(wsCopy)

# --- Sheet1 표 캡처 이미지 생성 ---
FONT_PATH = "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc"
f_bold = ImageFont.truetype(FONT_PATH, 15, index=2)
f_reg = ImageFont.truetype(FONT_PATH, 14, index=0)

def pct(v):
    return "" if v is None else f"{v*100:.2f}%"
_bs = [v for v in VAL_B.values() if v is not None]
_cs = [v for v in VAL_C.values() if v is not None]
sumB = sum(_bs) if _bs else None
sumC = sum(_cs) if _cs else None
rows_data = [
    ("", f"{sunMonth}월 1주차", f"{monMonth}월 누적") if isCrossMonth else (주차텍스트, "주", "월"),
    ("사업부", pct(sumB), pct(sumC)),
    ("가전", pct(VAL_B['가전']), pct(VAL_C['가전'])),
    ("리빙", pct(VAL_B['리빙']), pct(VAL_C['리빙'])),
    ("주방", pct(VAL_B['주방']), pct(VAL_C['주방'])),
]
name_colors = {2: (255, 255, 0), 3: (255, 192, 0), 4: (198, 224, 180)}

def draw_table(W, H, scale=3):
    """목표 크기(W x H)에 맞춰 표를 직접 그린다. 리사이즈 없음 → 찌부 없음."""
    w, h = W * scale, H * scale
    cw = [round(w * r) for r in (0.269, 0.3655, 0.3655)]
    cw[-1] = w - cw[0] - cw[1]
    rh = [h // 5] * 5
    rh[-1] = h - sum(rh[:4])
    fs = max(8, int(min(rh) * 0.52))          # 행 높이에 비례한 폰트 크기
    fb = ImageFont.truetype(FONT_PATH, fs, index=2)
    fr = ImageFont.truetype(FONT_PATH, fs, index=0)
    im = PILImage.new("RGB", (w, h), "white")
    dr = ImageDraw.Draw(im)
    y = 0
    for ri, rowd in enumerate(rows_data):
        x = 0
        for ci, text in enumerate(rowd):
            if ri == 0:
                fill = (221, 235, 247)
            elif ci == 0 and ri in name_colors:
                fill = name_colors[ri]
            else:
                fill = (255, 255, 255)
            dr.rectangle([x, y, x + cw[ci], y + rh[ri]], fill=fill,
                         outline=(0, 0, 0), width=scale)          # 표시 1px
            f = fb if ri in (0, 1) else fr
            bb = dr.textbbox((0, 0), text, font=f)
            dr.text((x + (cw[ci] - (bb[2]-bb[0])) / 2,
                     y + (rh[ri] - (bb[3]-bb[1])) / 2 - bb[1]),
                    text, fill="black", font=f)
            x += cw[ci]
        y += rh[ri]
    dr.rectangle([0, 0, w-1, h-1], outline=(0, 0, 0), width=scale*3)   # 표시 3px = 요일 구분선과 동일
    return im

# --- 헤더 블록(AL1:AQ5)에 딱 맞추기 ---
# 1) 블록 실제 폭 계산 (AL~AQ)
def col_px(ws, c):
    w = ws.column_dimensions[get_column_letter(c)].width
    return round(w * 7) + 5 if w else 64

BLOCK_W = sum(col_px(wsTV, c) for c in range(38, MAX_COL + 1))  # AL:AQ

# 2) 표 비율(3열x5행)에 맞는 블록 높이를 역산해서 1~5행 높이 재설정
INSET = 3
import os
ROW1_PT = float(os.environ.get("ROW1_PT", "24.75"))   # 맨 윗행 높이(pt)
for _ws in (wsTV, wsCopy):
    _ws.row_dimensions[1].height = ROW1_PT
    for _r in range(2, 6):
        _ws.row_dimensions[_r].height = 15             # 렌더러 자동확장 방지

BLOCK_H = sum(round((wsTV.row_dimensions[r].height) * 96 / 72) for r in range(1, 6))
target_img_w = BLOCK_W - INSET * 2
target_img_h = BLOCK_H - INSET * 2

img_hi = draw_table(target_img_w, target_img_h)       # 목표 비율로 직접 그림
img_path = os.path.join(os.path.dirname(os.path.abspath(OUT)) or ".", "_capture.png")
img_hi.save(img_path)

EMU = 9525
xl_img = XLImage(img_path)
xl_img.anchor = OneCellAnchor(
    _from=AnchorMarker(col=37, colOff=INSET*EMU, row=0, rowOff=INSET*EMU),
    ext=XDRPositiveSize2D(cx=target_img_w*EMU, cy=target_img_h*EMU),
)
wsTV.add_image(xl_img)
print(f"블록 {BLOCK_W}x{BLOCK_H}px / 이미지 {target_img_w}x{target_img_h}px / 1행 {ROW1_PT}pt")

# 병합범위 둘레 테두리 확정 (저장 직전) — 헤더 우측 외곽선 포함
finalize_merged_borders(wsTV, row_offset=2)
finalize_merged_borders(wsCopy, row_offset=2)

wb.save(OUT)

# ---------------------------------------------------------------
# [중요] openpyxl 저장 시 기본 스타일(Normal)의 폰트 참조가 엉뚱한 폰트로 바뀐다.
#  엑셀 열 너비 단위 = "기본 폰트 기준 글자 수" 이므로, 기본 폰트가 Arial 9 등으로
#  바뀌면 같은 width 값이어도 열이 좁게 렌더된다(1546px vs 1884px).
#  원본 RAW 는 모두 fontId=0(Calibri 11) 이므로 저장 후 되돌린다.
# ---------------------------------------------------------------
def fix_normal_font(path):
    """기본 스타일(Normal)의 폰트를 Calibri 12 로 고정.

    - openpyxl 저장 시 Normal 의 폰트 참조가 파일마다 제각각(Arial 9 / 이름없음 / Calibri 12)이 되고,
      엑셀 열 너비 단위 = "기본 폰트 기준 글자 수" 이므로 같은 width 라도 렌더 폭이 달라진다.
      (Arial 9 → 1546px, Calibri 11 → 1715px, Calibri 12 → 1869px)
    - 사용자 실제 엑셀 캡처 비율 1.74 에 가장 가까운 값이 Calibri 12(1.70) 이므로 이 값으로 통일.
      (LibreOffice 의 Carlito 대체 폰트가 엑셀 Calibri 보다 좁아서 생기는 보정치)
    - 셀은 모두 자체 fontId 를 가지므로 글자 모양에는 영향 없고 열 너비 환산에만 쓰인다.
    """
    import zipfile, shutil, re as _re
    zin = zipfile.ZipFile(path)
    items = {n: zin.read(n) for n in zin.namelist()}
    zin.close()
    st = items["xl/styles.xml"].decode("utf-8")

    NORMAL = '<font><sz val="12"/><name val="Calibri"/><family val="2"/></font>'
    mf = _re.search(r'<fonts count="(\d+)"([^>]*)>', st)
    if not mf:
        return None
    n_old = int(mf.group(1))
    st = st.replace(mf.group(0), f'<fonts count="{n_old+1}"{mf.group(2)}>', 1)
    st = st.replace("</fonts>", NORMAL + "</fonts>", 1)   # 새 폰트를 맨 끝에 추가

    mx = _re.search(r"<cellStyleXfs[^>]*>\s*<xf[^>]*/?>", st)
    if not mx:
        return None
    before = mx.group(0)
    if 'fontId="' in before:
        after = _re.sub(r'fontId="\d+"', f'fontId="{n_old}"', before, count=1)
    else:
        after = before.replace("<xf ", f'<xf fontId="{n_old}" ', 1)
    st = st.replace(before, after, 1)

    items["xl/styles.xml"] = st.encode("utf-8")
    tmp = path + ".tmp"
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for n, data in items.items():
            zout.writestr(n, data)
    shutil.move(tmp, path)
    old = _re.search(r'fontId="(\d+)"', before)
    return f"Normal 폰트 fontId {old.group(1) if old else '?'} → {n_old} (Calibri 12)"

_r = fix_normal_font(OUT)
if _r:
    print(f"기본 스타일 폰트 보정: {_r}")

print("saved:", OUT)

# 검증 출력
print("wsTV merges:", [str(m) for m in wsTV.merged_cells.ranges if m.min_row <= 8][:12])
print("wsTV AL1:", repr(wsTV["AL1"].value))
print("wsCopy AL1:", repr(wsCopy["AL1"].value))
