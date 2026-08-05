import os

import openpyxl
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.utils import get_column_letter


INFILE = os.environ.get("INFILE")

if not INFILE:
    raise SystemExit("INFILE 환경변수에 대상 xlsx 경로를 지정하세요")

workbook = openpyxl.load_workbook(INFILE)

for sheet_name, last_row in (
    ("(TV)편성기획", 32),
    ("(TV)편성기획-팀편성", 32),
):
    worksheet = workbook[sheet_name]

    width_pixels = sum(
        round(
            (
                worksheet.column_dimensions[
                    get_column_letter(column)
                ].width
                or 8.43
            )
            * 7
        )
        + 5
        for column in range(1, 44)
    )

    height_pixels = sum(
        round(
            (
                worksheet.row_dimensions[row].height
                or 15
            )
            * 96
            / 72
        )
        for row in range(1, last_row + 1)
    )

    # 표보다 살짝 큰 사용자 정의 용지를 사용해
    # 100% 배율을 유지하면서 가로·세로 중앙에 배치합니다.
    worksheet.page_setup.orientation = "portrait"
    worksheet.page_setup.fitToPage = False
    worksheet.page_setup.scale = 100

    worksheet.page_setup.paperWidth = (
        f"{width_pixels * 1.12 / 96 * 25.4:.2f}mm"
    )
    worksheet.page_setup.paperHeight = (
        f"{height_pixels * 1.12 / 96 * 25.4:.2f}mm"
    )

    worksheet.sheet_properties.pageSetUpPr = (
        PageSetupProperties(fitToPage=False)
    )

    worksheet.print_area = f"A1:AQ{last_row}"

    worksheet.page_margins.left = 0
    worksheet.page_margins.right = 0
    worksheet.page_margins.top = 0
    worksheet.page_margins.bottom = 0
    worksheet.page_margins.header = 0
    worksheet.page_margins.footer = 0

    # PDF 안에서 표를 용지의 가로·세로 중앙에 배치합니다.
    worksheet.print_options.horizontalCentered = True
    worksheet.print_options.verticalCentered = True

workbook["Sheet1"].sheet_state = "hidden"

output_file = os.environ.get(
    "OUTFILE",
    "render_exact.xlsx",
)

workbook.save(output_file)
print("ok")
