# -*- coding: utf-8 -*-
"""편성 비중 리포트(xlsx) 에서 가전/리빙/주방 비중 뽑기

회사 시스템에서 받은 비중 리포트 형식:
    A열 사업부 | B열 팀 | C열 파트 | D열 현재주차 | E열 마케팅 | F열 월누적 | ...

  리빙사업부 | 가전 |      | 4.1  | 0.0 | 4.51   ← 팀 레벨 (이 행을 씀)
  리빙사업부 |      | 가전 | 4.1  | 0.0 | 4.51   ← 파트 레벨 (중복이라 건너뜀)

D열(현재주차) → 뷰어의 "주",  F열(월누적) → 뷰어의 "월"
"""
import os
import openpyxl

DEPTS = ("가전", "리빙", "주방")
COL_TEAM, COL_WEEK, COL_MONTH = 2, 4, 6      # B, D, F


def _clean(v):
    return str(v).strip().strip("'").strip() if v not in (None, "") else ""


def read(path):
    """→ ({주 비중}, {월 비중})  / 못 읽으면 예외"""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    week, month = {}, {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=COL_MONTH):
        team = _clean(row[COL_TEAM - 1].value)
        if team not in DEPTS or team in week:
            continue
        w = row[COL_WEEK - 1].value
        m = row[COL_MONTH - 1].value
        try:
            week[team] = float(w) / 100
            month[team] = float(m) / 100
        except (TypeError, ValueError):
            raise ValueError(f"{os.path.basename(path)}: '{team}' 행의 값이 숫자가 아닙니다 ({w!r}, {m!r})")
    wb.close()

    missing = [d for d in DEPTS if d not in week]
    if missing:
        raise ValueError(f"{os.path.basename(path)}: {missing} 를 찾지 못했습니다 "
                         f"(B열 '팀' 에 가전/리빙/주방 이 있어야 합니다)")
    return week, month


if __name__ == "__main__":
    import sys
    w, m = read(sys.argv[1])
    print("주   :", {k: f"{v*100:.2f}%" for k, v in w.items()}, f"합 {sum(w.values())*100:.2f}%")
    print("월누적:", {k: f"{v*100:.2f}%" for k, v in m.items()}, f"합 {sum(m.values())*100:.2f}%")
